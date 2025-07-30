import tkinter as tk
from tkinter import filedialog
import os
from tkinter import ttk
import time
import subprocess
import threading
import openpyxl
from tkinter import messagebox
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill

# Create the main application window
root = tk.Tk()
root.title("Folder Selector")
root.geometry("600x375")

selected_folder = tk.StringVar()
selected_folder.set("No folder selected.")

file_paths = []

status_text = tk.StringVar()
status_text.set("")
progress_var = tk.DoubleVar()

network_drive_path = tk.StringVar()
network_drive_path.set("")

current_folder_text = tk.StringVar()
current_folder_text.set("")

loading_text = tk.StringVar()
loading_label = tk.Label(root, textvariable=loading_text, fg='blue', font=(None, 12, 'bold'))
# Remove the initial pack for loading_label here

# Animate the loading label
def animate_loading():
    if not getattr(animate_loading, 'running', False):
        return
    dots = animate_loading.dots = (getattr(animate_loading, 'dots', 0) + 1) % 8
    loading_text.set(f"Counting files\n" + " " + "." * dots)
    root.after(400, animate_loading)

# Animate the saving label
def animate_saving():
    if not getattr(animate_saving, 'running', False):
        return
    dots = animate_saving.dots = (getattr(animate_saving, 'dots', 0) + 1) % 8
    saving_text.set(f"Saving file\n" + " " + "." * dots)
    root.after(400, animate_saving)

def fast_count_files(folder):
    result = subprocess.run(
        ["powershell", "-Command", f"(Get-ChildItem -Path '{folder}' -Recurse -File | Measure-Object).Count"],
        capture_output=True, text=True
    )
    try:
        return int(result.stdout.strip())
    except Exception:
        return None

def select_folder():
    global file_paths
    folder = filedialog.askdirectory()
    if not folder:
        selected_folder.set("No folder selected.")
        file_paths = []
        current_folder_text.set("")
        status_text.set("")
        progress_var.set(0)
        loading_text.set("")
        return
    selected_folder.set(f"Selected Folder: {folder}")
    loading_text.set("Counting files")
    root.update_idletasks()
    animate_loading.running = True
    animate_loading.dots = 0
    animate_loading()

    def worker():
        nonlocal folder
        # Count total files first (Python version, with live updates)
        total_files = 0
        for root_dir, _, files in os.walk(folder):
            rel_path = os.path.relpath(root_dir, folder)
            if rel_path == ".":
                display_path = "(root)"
            else:
                display_path = rel_path
            root.after(0, lambda p=display_path: current_folder_text.set(f"Scanning Files in: {p}"))
            root.after(0, root.update_idletasks)
            total_files += len(files)
        def remove_loading():
            loading_text.set("")
            animate_loading.running = False
        root.after(0, remove_loading)
        file_paths_local = []
        current_file_num = 0
        for root_dir, _, files in os.walk(folder):
            rel_path = os.path.relpath(root_dir, folder)
            if rel_path == ".":
                display_path = "(root)"
            else:
                display_path = rel_path
            root.after(0, lambda p=display_path: current_folder_text.set(f"Searching Subfolder: {p}"))
            for file in files:
                file_path = os.path.join(root_dir, file)
                file_paths_local.append(file_path)
                current_file_num += 1
                percent = (current_file_num / total_files) * 100 if total_files else 0
                root.after(0, lambda f=file, n=current_file_num, t=total_files, p=percent: (
                    status_text.set(f"Adding: {f} ({n}/{t})"),
                    progress_var.set(p)
                ))
                root.after(0, root.update_idletasks)
                # print("Added:", file_path)
                time.sleep(0.03)
        def finish():
            current_folder_text.set("")
            status_text.set(f"Done! {total_files} files added.")
            progress_var.set(100)
            global file_paths
            file_paths = file_paths_local
            # print("All file paths:", file_paths)
        root.after(0, finish)
    threading.Thread(target=worker, daemon=True).start()

# Button to open the folder selection dialog
select_button = tk.Button(root, text="Select Folder", command=select_folder, font=(None, 9), height=1, width=20)
select_button.pack(padx=10, pady=(10, 2), anchor="center")

# Label to display the selected folder
folder_label = tk.Label(root, textvariable=selected_folder, wraplength=380, anchor="w", justify="left")
folder_label.pack(fill="x", padx=10, pady=(0, 10))

# Label to display the current folder being searched
current_folder_label = tk.Label(root, textvariable=current_folder_text, anchor="w", justify="left")
current_folder_label.pack(fill="x", padx=10, pady=(0, 2))

# Status bar (label)
status_label = tk.Label(root, textvariable=status_text, anchor="w", justify="left")
status_label.pack(fill="x", padx=10, pady=(0, 2))

# Progress bar at the bottom
progress_bar = ttk.Progressbar(root, variable=progress_var, maximum=100)
progress_bar.pack(fill="x", padx=10, pady=(0, 10))

# Loading label (Counting files...) at the bottom, just above network drive controls
loading_label.pack(fill="x", padx=10, pady=(0, 2))

# Function to generate Excel file from file_paths

def generate_excel_file():
    if not file_paths:
        messagebox.showwarning("No Files", "No files to write. Please select a folder and scan files first.")
        return
    # Parse selected network drive mapping
    drive_letter = None
    unc_path = None
    ndp = network_drive_path.get()
    if ndp.startswith("Selected Network Drive: "):
        try:
            # Format: Selected Network Drive: Z: (\\server\share)
            main = ndp[len("Selected Network Drive: "):]
            drive_letter, unc = main.split(" (")
            drive_letter = drive_letter.strip()
            unc_path = unc.rstrip(")").strip()
        except Exception:
            pass
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "File Paths"
    ws.cell(row=1, column=1, value="File Path")
    ws.cell(row=1, column=2, value="Length")
    for idx, path in enumerate(file_paths, start=2):
        norm_path = os.path.normpath(path)
        if drive_letter and unc_path and norm_path.startswith(drive_letter + "\\"):
            new_path = unc_path + norm_path[len(drive_letter):]
        else:
            new_path = norm_path
        ws.cell(row=idx, column=1, value=new_path)
        ws.cell(row=idx, column=2, value=len(new_path))
    # Add conditional formatting for length >= 260
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    ws.conditional_formatting.add(
        f'B2:B{idx}',
        CellIsRule(operator='greaterThanOrEqual', formula=['260'], stopIfTrue=True, fill=red_fill, font=red_font)
    )
    try:
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Save Excel File")
        if save_path:
            # Start saving animation
            animate_saving.running = True
            animate_saving.dots = 0
            animate_saving()
            root.update_idletasks()
            wb.save(save_path)
            # Stop animation and clear label
            animate_saving.running = False
            saving_text.set("File saved!")
            messagebox.showinfo("Success", f"Excel file saved to:\n{save_path}")
    except Exception as e:
        animate_saving.running = False
        saving_text.set("")
        messagebox.showerror("Error", f"Failed to save Excel file:\n{e}")

# Button to generate Excel file at the very bottom
excel_button = tk.Button(root, text="Generate Excel File", command=generate_excel_file, font=(None, 9), height=1, width=20)
excel_button.pack(padx=10, pady=(10, 10), anchor="center", side="bottom")

saving_text = tk.StringVar()
saving_label = tk.Label(root, textvariable=saving_text, fg='blue', font=(None, 12, 'bold'))
saving_label.pack(fill="x", padx=10, pady=(0, 2), side="bottom")

def select_network_drive():
    # Run 'net use' to get mapped network drives
    try:
        result = subprocess.run(["powershell", "-Command", "net use"], capture_output=True, text=True, check=True)
        output = result.stdout
        # Parse lines that look like: X:        \\server\share
        drives = []
        for line in output.splitlines():
            if ":" in line and "\\" in line:
                parts = line.split()
                if len(parts) >= 2 and parts[0].endswith(":"):
                    drives.append((parts[0], parts[1]))
        if not drives:
            network_drive_path.set("No network drives found.")
            return
        elif len(drives) == 1:
            selected = drives[0]
        else:
            # If multiple, let user select
            import tkinter.simpledialog
            options = [f"{d[0]} ({d[1]})" for d in drives]
            choice = tkinter.simpledialog.askstring("Select Network Drive", "Enter drive letter (e.g. Z:):\n" + "\n".join(options))
            selected = next((d for d in drives if d[0].lower() == (choice or '').strip().lower()), None)
            if not selected:
                network_drive_path.set("No network drive selected.")
                return
        network_drive_path.set(f"Selected Network Drive: {selected[0]} ({selected[1]})")
    except Exception as e:
        network_drive_path.set(f"Error: {e}")

# Button to select network drive
network_drive_button = tk.Button(root, text="Select Network Drive", command=select_network_drive, font=(None, 9), height=1, width=20)
network_drive_button.pack(padx=10, pady=(10, 2), anchor="center")

# Label to display selected network drive
network_drive_label = tk.Label(root, textvariable=network_drive_path, anchor="w", justify="left")
network_drive_label.pack(fill="x", padx=10, pady=(0, 10))


root.mainloop()
